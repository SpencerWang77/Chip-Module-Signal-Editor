"""Workbook sheet-discovery, preservation, and upload-rule UI tests."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from PyQt5.QtCore import QEvent
from PyQt5.QtWidgets import QApplication, QLabel

from chip_editor.ui.import_rules import ImportRulesCard
from chip_editor.workbook_io import (
    WorkbookFormatError,
    export_workbook,
    parse_register_workbook,
)


PROJECT_DIR = Path(__file__).resolve().parents[1]
SAMPLE_WORKBOOK = PROJECT_DIR / "top_signal.xlsx"


def _sheet_snapshot(worksheet) -> dict[str, object]:
    return {
        "values": tuple(
            tuple(cell.value for cell in row)
            for row in worksheet.iter_rows()
        ),
        "merged": tuple(sorted(str(item) for item in worksheet.merged_cells.ranges)),
        "state": worksheet.sheet_state,
        "freeze_panes": str(worksheet.freeze_panes or ""),
    }


class WorkbookContractTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.temporary_directory = tempfile.TemporaryDirectory()
        cls.renamed_workbook = (
            Path(cls.temporary_directory.name) / "renamed-sheets.xlsx"
        )

        workbook = load_workbook(SAMPLE_WORKBOOK, data_only=False)
        workbook["3.register"].title = "Customer REGISTER Map"
        workbook["2.AD_AA"].title = "Analog AD_AA Reference"
        readme = workbook.create_sheet("register README", 0)
        readme["A1"] = "This matching sheet name is intentionally not a register table."
        notes = workbook.create_sheet("Keep Everything")
        notes["A1"] = "Preserve this text"
        notes["B2"] = "=21*2"
        notes.merge_cells("C3:D4")
        notes["C3"] = "Preserve this merge"
        notes.freeze_panes = "B2"
        notes.sheet_state = "hidden"
        workbook.save(cls.renamed_workbook)

    @classmethod
    def tearDownClass(cls) -> None:
        cls.temporary_directory.cleanup()

    def test_sheet_names_are_discovered_by_case_insensitive_substring(self) -> None:
        data = parse_register_workbook(self.renamed_workbook)
        self.assertEqual(data.sheet_name, "Customer REGISTER Map")
        self.assertEqual(data.description_sheet_name, "Analog AD_AA Reference")
        self.assertGreater(data.described_field_count, 0)
        described_field = next(
            field
            for register in data.registers
            for field in register.fields
            if field.has_description
        )
        self.assertEqual(
            described_field.description_source_sheet,
            "Analog AD_AA Reference",
        )

    def test_ad_aa_sheet_is_optional(self) -> None:
        path = Path(self.temporary_directory.name) / "without-descriptions.xlsx"
        workbook = load_workbook(self.renamed_workbook, data_only=False)
        workbook.remove(workbook["Analog AD_AA Reference"])
        workbook.save(path)

        data = parse_register_workbook(path)
        self.assertEqual(data.sheet_name, "Customer REGISTER Map")
        self.assertEqual(data.description_sheet_name, "")
        self.assertEqual(data.described_field_count, 0)

    def test_present_but_unreadable_ad_aa_sheet_is_rejected(self) -> None:
        path = Path(self.temporary_directory.name) / "bad-descriptions.xlsx"
        workbook = load_workbook(self.renamed_workbook, data_only=False)
        worksheet = workbook["Analog AD_AA Reference"]
        worksheet["D1"] = "NOT_A_SIGNAL_HEADER"
        worksheet["Q1"] = "NOT_A_DESCRIPTION_HEADER"
        workbook.save(path)

        with self.assertRaisesRegex(WorkbookFormatError, "AD_AA.*not readable"):
            parse_register_workbook(path)

    def test_missing_register_named_sheet_is_rejected(self) -> None:
        path = Path(self.temporary_directory.name) / "no-register-name.xlsx"
        workbook = load_workbook(SAMPLE_WORKBOOK, data_only=False)
        workbook["3.register"].title = "Signal Map"
        workbook.save(path)

        with self.assertRaisesRegex(WorkbookFormatError, "contains “register”"):
            parse_register_workbook(path)

    def test_export_changes_register_sheet_and_preserves_every_other_sheet(self) -> None:
        data = parse_register_workbook(self.renamed_workbook)
        register = next(
            item
            for item in data.registers
            if any(field.original_name == "pll_inst_reg1" for field in item.fields)
        )
        register_field = next(
            field
            for field in register.fields
            if field.original_name == "pll_inst_reg1"
        )
        register_field.name = "pll_inst_vco_control"
        register.bits[0] = 1 - register.bits[0]

        destination = Path(self.temporary_directory.name) / "preserved-export.xlsx"
        export_workbook(data, destination)

        source = load_workbook(self.renamed_workbook, data_only=False)
        exported = load_workbook(destination, data_only=False)
        self.assertEqual(exported.sheetnames, source.sheetnames)
        for name in source.sheetnames:
            if name == data.sheet_name:
                continue
            self.assertEqual(
                _sheet_snapshot(exported[name]),
                _sheet_snapshot(source[name]),
                name,
            )

        exported_field = exported[data.sheet_name].cell(
            register.worksheet_row,
            register_field.start_column,
        )
        source_field = source[data.sheet_name].cell(
            register.worksheet_row,
            register_field.start_column,
        )
        self.assertTrue(str(exported_field.value).startswith("pll_inst_vco_control"))
        self.assertFalse(str(source_field.value).startswith("pll_inst_vco_control"))
        output_headers = [
            cell.value for cell in exported[data.sheet_name][data.header_row]
        ]
        self.assertIn("EDITED_HEX_VALUE", output_headers)


class ImportRulesUiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.application = QApplication.instance() or QApplication([])

    def test_info_card_expands_and_explains_the_full_contract(self) -> None:
        card = ImportRulesCard()
        self.assertTrue(card.details.isHidden())
        card.show()
        self.application.sendEvent(card.info_button, QEvent(QEvent.Enter))
        self.application.processEvents()
        self.assertTrue(card.expanded)
        self.assertFalse(card.details.isHidden())
        text = " ".join(
            label.text()
            for label in card.findChildren(QLabel)
        )
        self.assertIn("contains “register”", text)
        self.assertIn("contains “AD_AA”", text)
        self.assertIn("every other worksheet", text)
        card.set_expanded(False)
        self.assertTrue(card.details.isHidden())
        card.deleteLater()


if __name__ == "__main__":
    unittest.main()
