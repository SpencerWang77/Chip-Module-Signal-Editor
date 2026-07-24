"""Description matching and selected-field editing tests."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from PyQt5.QtWidgets import QApplication, QLabel, QLineEdit

from chip_editor.models import build_register_modules
from chip_editor.ui.editor_page import EditorPage
from chip_editor.workbook_io import export_workbook, parse_register_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
SAMPLE_WORKBOOK = PROJECT_DIR / "top_signal.xlsx"


def _field(data, name: str):
    for register in data.registers:
        for register_field in register.fields:
            if register_field.original_name == name:
                return register, register_field
    raise AssertionError(f"Could not find field {name}")


class DescriptionParsingTests(unittest.TestCase):
    def setUp(self) -> None:
        self.data = parse_register_workbook(SAMPLE_WORKBOOK)

    def test_address_aware_suffix_match(self) -> None:
        register, register_field = _field(self.data, "pll_inst_reg1")
        self.assertEqual(register.hex_addr, "0x0018")
        self.assertEqual(register_field.description_source_name, "reg1[4:0]")
        self.assertIn("PLL-VCO", register_field.description)

    def test_numbered_byte_chunks_share_wide_signal_description(self) -> None:
        _, low_chunk = _field(self.data, "adc_inst_BOUTL_0")
        _, high_chunk = _field(self.data, "adc_inst_BOUTL_1")
        self.assertEqual(low_chunk.description_source_name, "BOUTL[11:0]")
        self.assertEqual(high_chunk.description, low_chunk.description)

    def test_short_global_suffix_is_not_mismatched(self) -> None:
        _, exact_address_en = _field(self.data, "osc32m_inst_EN")
        _, unrelated_test_en = _field(self.data, "pll_inst_test_en")
        self.assertTrue(exact_address_en.has_description)
        self.assertFalse(unrelated_test_en.has_description)

    def test_renamed_field_is_exported_to_register_sheet(self) -> None:
        register, register_field = _field(self.data, "pll_inst_reg1")
        register_field.name = "pll_inst_vco_control"
        register_field.description = "Updated PLL VCO control description."
        with tempfile.TemporaryDirectory() as temporary_directory:
            destination = Path(temporary_directory) / "edited.xlsx"
            export_workbook(self.data, destination)
            workbook = load_workbook(destination, data_only=False)
            cell = workbook[self.data.sheet_name].cell(
                register.worksheet_row,
                register_field.start_column,
            )
            self.assertTrue(str(cell.value).startswith("pll_inst_vco_control"))
            description_cell = workbook[
                register_field.description_source_sheet
            ].cell(
                register_field.description_source_row,
                register_field.description_source_column,
            )
            self.assertEqual(
                description_cell.value,
                "Updated PLL VCO control description.",
            )


class FieldDetailsUiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.application = QApplication.instance() or QApplication([])

    def test_details_panel_rename_updates_inline_model_and_log(self) -> None:
        data = parse_register_workbook(SAMPLE_WORKBOOK)
        page = EditorPage()
        page.load_data(data)
        module = next(
            module
            for module in build_register_modules(data.registers)
            if module.name == "PLL_WADDR_0"
        )
        page.open_module(module)
        register, register_field = _field(data, "pll_inst_reg1")

        page.show_field_details(register, register_field)
        self.assertIn(
            "PLL-VCO",
            page.field_details.description.toPlainText(),
        )
        page.field_details.name_edit.setText("pll_inst_vco_control")
        page.field_details._request_rename()

        self.assertEqual(register_field.name, "pll_inst_vco_control")
        self.assertEqual(page.change_log.count, 1)
        row = page.row_by_register_id[id(register)]
        inline_names = [
            label.text() for _, label in row.bit_grid.field_labels
        ]
        self.assertIn("pll_inst_vco_control", inline_names)
        self.assertTrue(
            all(
                isinstance(label, QLabel) and not isinstance(label, QLineEdit)
                for _, label in row.bit_grid.field_labels
            )
        )

        page.field_details.description.setPlainText(
            "Updated from the selected field panel."
        )
        page.field_details._request_description_change()
        self.assertEqual(
            register_field.description,
            "Updated from the selected field panel.",
        )
        self.assertEqual(page.change_log.count, 2)
        page.deleteLater()

    def test_shared_ad_aa_description_updates_every_matching_chunk(self) -> None:
        data = parse_register_workbook(SAMPLE_WORKBOOK)
        page = EditorPage()
        page.load_data(data)
        first_register, first_chunk = _field(data, "adc_inst_BOUTL_0")
        _, second_chunk = _field(data, "adc_inst_BOUTL_1")

        page.show_field_details(first_register, first_chunk)
        page.field_details.description.setPlainText(
            "Shared ADC output description."
        )
        page.field_details._request_description_change()

        self.assertEqual(
            first_chunk.description,
            "Shared ADC output description.",
        )
        self.assertEqual(second_chunk.description, first_chunk.description)
        self.assertEqual(page.change_log.count, 1)
        page.deleteLater()


if __name__ == "__main__":
    unittest.main()
