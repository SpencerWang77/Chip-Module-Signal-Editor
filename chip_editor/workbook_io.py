"""Excel validation, parsing, and export for register maps."""

from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from .constants import (
    BIT_COLUMN_COUNT,
    DESCRIPTION_SHEET_KEYWORD,
    REGISTER_SHEET_KEYWORD,
)
from .models import RegisterByte, RegisterField, WorkbookData


FIELD_VALUE_RE = re.compile(r"^(.*?)(?:\[\s*(0x[0-9a-fA-F]+|\d+)\s*\])?\s*$")
SIGNAL_RANGE_RE = re.compile(r"\s*\[[^\]]+\]\s*$")
FIELD_CHUNK_RE = re.compile(r"_\d+$")
HEX_ADDRESS_RE = re.compile(r"0[xX]([0-9a-fA-F]+)")


@dataclass(frozen=True)
class _DescriptionRecord:
    """One signal description imported from the optional AD_AA sheet."""

    worksheet_row: int
    signal_name: str
    normalized_suffix: str
    description: str
    description_column: int
    addresses: tuple[str, ...]


class WorkbookFormatError(ValueError):
    """Raised when an uploaded workbook does not match the register format."""


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _matching_sheet_names(workbook, keyword: str) -> list[str]:
    """Return worksheet names containing a keyword, ignoring letter case."""

    normalized_keyword = keyword.casefold()
    return [
        name
        for name in workbook.sheetnames
        if normalized_keyword in name.casefold()
    ]


def _find_header(
    worksheet,
    required_headers: set[str],
) -> tuple[Optional[int], dict[str, int]]:
    """Find a table header within the first 40 rows, ignoring header case."""

    for row_index in range(1, min(worksheet.max_row, 40) + 1):
        found = {
            _cell_text(worksheet.cell(row_index, column).value).upper(): column
            for column in range(1, worksheet.max_column + 1)
        }
        if required_headers.issubset(found):
            return row_index, found
    return None, {}


def _find_register_sheet(workbook):
    """Choose the first correctly formatted sheet whose name contains register."""

    candidate_names = _matching_sheet_names(workbook, REGISTER_SHEET_KEYWORD)
    if not candidate_names:
        raise WorkbookFormatError(
            "The workbook needs a worksheet whose name contains “register” "
            f"(case-insensitive). Found: {', '.join(workbook.sheetnames)}"
        )

    required_headers = {"HEX_BYTE_ADDR", "REG_NAME", "REG_FIELD"}
    for name in candidate_names:
        worksheet = workbook[name]
        header_row, header_columns = _find_header(worksheet, required_headers)
        if header_row is not None:
            return worksheet, header_row, header_columns

    raise WorkbookFormatError(
        "Found a worksheet whose name contains “register”, but it is not readable. "
        "It must include HEX_BYTE_ADDR, REG_NAME, and REG_FIELD headers within "
        f"the first 40 rows. Checked: {', '.join(candidate_names)}"
    )


def _hex_address(value: object, decimal_fallback: object = None) -> str:
    candidate = value if value is not None else decimal_fallback
    if candidate is None:
        return "—"
    if isinstance(candidate, int):
        return f"0x{candidate:04X}"
    text = str(candidate).strip()
    try:
        return f"0x{int(text, 0):04X}"
    except (TypeError, ValueError):
        return text


def _parse_field_text(value: object) -> tuple[str, int]:
    text = _cell_text(value)
    if not text:
        return "", 0
    match = FIELD_VALUE_RE.match(text)
    if not match:
        return text, 0
    name = match.group(1).strip()
    raw_value = match.group(2)
    try:
        default = int(raw_value, 0) if raw_value else 0
    except ValueError:
        default = 0
    return name, default


def _description_addresses(value: object) -> tuple[str, ...]:
    """Normalize one or several semicolon/newline-separated byte addresses."""

    if value is None:
        return ()
    if isinstance(value, int):
        return (_hex_address(value),)
    return tuple(
        f"0x{int(match, 16):04X}"
        for match in HEX_ADDRESS_RE.findall(str(value))
    )


def _description_records(workbook) -> tuple[str, list[_DescriptionRecord]]:
    """Read the optional signal-description table without making it mandatory."""

    candidate_names = _matching_sheet_names(workbook, DESCRIPTION_SHEET_KEYWORD)
    if not candidate_names:
        return "", []

    required_headers = {"SIGNAL_NAME", "DESCRIPTION"}
    worksheet = None
    header_row: Optional[int] = None
    header_columns: dict[str, int] = {}
    for name in candidate_names:
        candidate = workbook[name]
        candidate_header_row, candidate_header_columns = _find_header(
            candidate,
            required_headers,
        )
        if candidate_header_row is not None:
            worksheet = candidate
            header_row = candidate_header_row
            header_columns = candidate_header_columns
            break

    if worksheet is None or header_row is None:
        raise WorkbookFormatError(
            "Found a worksheet whose name contains “AD_AA”, but it is not readable. "
            "It must include SIGNAL_NAME and DESCRIPTION headers within the first "
            f"40 rows. Checked: {', '.join(candidate_names)}"
        )

    signal_column = header_columns["SIGNAL_NAME"]
    description_column = header_columns["DESCRIPTION"]
    address_column = header_columns.get("HEX_BYTE_ADDR")
    records: list[_DescriptionRecord] = []
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        signal_name = _cell_text(worksheet.cell(row_index, signal_column).value)
        description = _cell_text(worksheet.cell(row_index, description_column).value)
        if not signal_name or not description:
            continue
        normalized_suffix = SIGNAL_RANGE_RE.sub("", signal_name).strip().casefold()
        if not normalized_suffix:
            continue
        address_value = (
            worksheet.cell(row_index, address_column).value if address_column else None
        )
        records.append(
            _DescriptionRecord(
                worksheet_row=row_index,
                signal_name=signal_name,
                normalized_suffix=normalized_suffix,
                description=description,
                description_column=description_column,
                addresses=_description_addresses(address_value),
            )
        )
    return worksheet.title, records


def _suffix_matches(field_target: str, description_suffix: str) -> bool:
    """Require a complete underscore/dot-delimited suffix, not a substring."""

    if field_target == description_suffix:
        return True
    if not field_target.endswith(description_suffix):
        return False
    boundary_index = len(field_target) - len(description_suffix) - 1
    return boundary_index >= 0 and field_target[boundary_index] in "_."


def _match_description(
    field_name: str,
    register_address: str,
    records: list[_DescriptionRecord],
) -> Optional[_DescriptionRecord]:
    """Find the most specific, address-aware suffix match for a register field."""

    normalized_field = field_name.strip().casefold()
    chunkless_field = FIELD_CHUNK_RE.sub("", normalized_field)
    targets = [(normalized_field, False)]
    if chunkless_field != normalized_field:
        targets.append((chunkless_field, True))

    candidates: list[tuple[_DescriptionRecord, bool]] = []
    for record in records:
        for target, chunk_removed in targets:
            if _suffix_matches(target, record.normalized_suffix):
                candidates.append((record, chunk_removed))
                break

    if not candidates:
        return None

    address_matches = [
        candidate
        for candidate in candidates
        if register_address in candidate[0].addresses
    ]
    if address_matches:
        candidates = address_matches
    else:
        # Very short global suffixes such as EN are too ambiguous. A short suffix
        # remains safe for explicitly numbered chunks such as gpio_dr_0 → DR.
        candidates = [
            candidate
            for candidate in candidates
            if len(candidate[0].normalized_suffix) >= 3 or candidate[1]
        ]
        if not candidates:
            return None

    longest_suffix = max(
        len(candidate[0].normalized_suffix) for candidate in candidates
    )
    candidates = [
        candidate
        for candidate in candidates
        if len(candidate[0].normalized_suffix) == longest_suffix
    ]

    # If equally specific global candidates disagree, omit the description rather
    # than displaying documentation for the wrong field.
    distinct_meanings = {
        (candidate[0].normalized_suffix, candidate[0].description)
        for candidate in candidates
    }
    if len(distinct_meanings) > 1:
        return None
    return min(candidates, key=lambda candidate: candidate[0].worksheet_row)[0]


def parse_register_workbook(path: Path) -> WorkbookData:
    """Validate and parse a workbook without changing the source file."""

    try:
        # Cached values resolve formula-based addresses in the supplied sample.
        # Export reopens with data_only=False so source formulas remain intact.
        workbook = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        raise WorkbookFormatError(f"This file could not be opened as an Excel workbook: {exc}") from exc

    worksheet, header_row, header_columns = _find_register_sheet(workbook)
    description_sheet_name, description_records = _description_records(workbook)

    bit_start = header_columns["REG_FIELD"]
    bit_end = bit_start + BIT_COLUMN_COUNT - 1
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= header_row <= merged_range.max_row
            and merged_range.min_col <= bit_start <= merged_range.max_col
        ):
            if merged_range.max_col - merged_range.min_col + 1 == BIT_COLUMN_COUNT:
                bit_start = merged_range.min_col
                bit_end = merged_range.max_col
            break

    if bit_end > worksheet.max_column:
        raise WorkbookFormatError("REG_FIELD must span eight columns, ordered from bit 7 to bit 0.")

    merged_by_row: dict[int, list[tuple[int, int]]] = {}
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.max_col < bit_start or merged_range.min_col > bit_end:
            continue
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            if row_index > header_row:
                merged_by_row.setdefault(row_index, []).append(
                    (max(bit_start, merged_range.min_col), min(bit_end, merged_range.max_col))
                )

    dec_col = header_columns.get("DEC_BYTE_ADDR")
    reg32_col = header_columns.get("REG32_NAME")
    hex_col = header_columns["HEX_BYTE_ADDR"]
    name_col = header_columns["REG_NAME"]
    registers: list[RegisterByte] = []

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        address = worksheet.cell(row_index, hex_col).value
        register_name = worksheet.cell(row_index, name_col).value
        decimal_address = worksheet.cell(row_index, dec_col).value if dec_col else None
        if address is None and decimal_address is None and register_name is None:
            continue

        fields: list[RegisterField] = []
        visited: set[int] = set()
        merged_spans = merged_by_row.get(row_index, [])

        for column in range(bit_start, bit_end + 1):
            if column in visited:
                continue
            span_start = column
            span_end = column
            for merged_start, merged_end in merged_spans:
                if merged_start <= column <= merged_end:
                    span_start, span_end = merged_start, merged_end
                    break
            visited.update(range(span_start, span_end + 1))

            cell = worksheet.cell(row_index, span_start)
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            original_text = _cell_text(cell.value)
            field_name, default_value = _parse_field_text(original_text)
            if not field_name:
                continue
            start_bit = 7 - (span_start - bit_start)
            end_bit = 7 - (span_end - bit_start)
            width = start_bit - end_bit + 1
            description_record = _match_description(
                field_name,
                _hex_address(address, decimal_address),
                description_records,
            )
            fields.append(
                RegisterField(
                    name=field_name,
                    original_name=field_name,
                    original_text=original_text,
                    start_column=span_start,
                    end_column=span_end,
                    start_bit=start_bit,
                    end_bit=end_bit,
                    default_value=default_value & ((1 << width) - 1),
                    description=(
                        description_record.description if description_record else ""
                    ),
                    original_description=(
                        description_record.description if description_record else ""
                    ),
                    description_source_name=(
                        description_record.signal_name if description_record else ""
                    ),
                    description_source_sheet=description_sheet_name,
                    description_source_row=(
                        description_record.worksheet_row if description_record else 0
                    ),
                    description_source_column=(
                        description_record.description_column
                        if description_record
                        else 0
                    ),
                )
            )

        bits = [0] * BIT_COLUMN_COUNT
        for register_field in fields:
            for bit_number in range(register_field.end_bit, register_field.start_bit + 1):
                display_index = 7 - bit_number
                bits[display_index] = (
                    register_field.default_value >> (bit_number - register_field.end_bit)
                ) & 1

        registers.append(
            RegisterByte(
                worksheet_row=row_index,
                dec_addr=_cell_text(decimal_address),
                hex_addr=_hex_address(address, decimal_address),
                reg32_name=_cell_text(worksheet.cell(row_index, reg32_col).value) if reg32_col else "",
                reg_name=_cell_text(register_name) or "Unnamed register",
                fields=fields,
                original_bits=tuple(bits),
            )
        )

    if not registers:
        raise WorkbookFormatError("No register rows were found below the header.")

    return WorkbookData(
        source_path=path,
        sheet_name=worksheet.title,
        header_row=header_row,
        bit_start_column=bit_start,
        registers=registers,
        description_sheet_name=description_sheet_name,
    )


def export_workbook(data: WorkbookData, destination: Path) -> None:
    """Export a copy with edited register fields and matched descriptions."""

    if data.source_path.resolve() == destination.resolve():
        raise ValueError("Choose a new file name so the source workbook remains unchanged.")

    shutil.copy2(data.source_path, destination)
    workbook = load_workbook(destination, data_only=False, read_only=False)
    worksheet = workbook[data.sheet_name]
    output_column = max(worksheet.max_column + 1, data.bit_start_column + BIT_COLUMN_COUNT)
    worksheet.cell(data.header_row, output_column, "EDITED_HEX_VALUE")
    description_updates: dict[tuple[str, int, int], str] = {}

    for register in data.registers:
        worksheet.cell(register.worksheet_row, output_column, f"0x{register.value:02X}")
        for register_field in register.fields:
            display_start = 7 - register_field.start_bit
            display_end = 7 - register_field.end_bit
            segment = register.bits[display_start : display_end + 1]
            original_segment = register.original_bits[display_start : display_end + 1]
            segment_changed = tuple(segment) != tuple(original_segment)
            name_changed = register_field.name != register_field.original_name
            if register_field.description_is_modified:
                description_key = (
                    register_field.description_source_sheet,
                    register_field.description_source_row,
                    register_field.description_source_column,
                )
                if all(description_key):
                    existing_update = description_updates.get(description_key)
                    if (
                        existing_update is not None
                        and existing_update != register_field.description
                    ):
                        raise ValueError(
                            "Conflicting edits target the same AD_AA description cell."
                        )
                    description_updates[description_key] = register_field.description
            if not segment_changed and not name_changed:
                continue
            new_value = RegisterByte.value_from_bits(segment)
            if segment_changed:
                output_text = f"{register_field.name}[0x{new_value:X}]"
            else:
                original_suffix = register_field.original_text[len(register_field.original_name) :]
                output_text = f"{register_field.name}{original_suffix}"
            worksheet.cell(
                register.worksheet_row,
                register_field.start_column,
                output_text,
            )

    for (sheet_name, row, column), description in description_updates.items():
        workbook[sheet_name].cell(row, column, description)

    workbook.save(destination)
